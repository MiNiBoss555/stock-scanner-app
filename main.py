import csv
import io
import json
import os
import hashlib
import hmac
import logging
import secrets
import sqlite3
import urllib.error
import urllib.request
from datetime import datetime, timedelta, timezone
from enum import Enum
from pathlib import Path
from typing import Any, Literal
from uuid import uuid4

from fastapi import FastAPI, File, Form, HTTPException, Query, Request, UploadFile, WebSocket, WebSocketDisconnect
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import Response
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field


class MovementType(str, Enum):
    IN = "in"
    OUT = "out"
    ISSUE = "issue"


class OrderStatus(str, Enum):
    NEW = "new"
    ASSIGNED = "assigned"
    PREPARING = "preparing"
    OUT_FOR_DELIVERY = "out_for_delivery"
    DELIVERED = "delivered"
    CANCELLED = "cancelled"


class Product(BaseModel):
    barcode: str = Field(..., min_length=3, description="Barcode or SKU")
    sku: str | None = None
    name: str
    unit: str = "pcs"
    minimum_stock: int = 0
    current_stock: int = 0
    category: str | None = None
    location: str | None = None


class User(BaseModel):
    user_id: str = Field(..., min_length=1)
    user_name: str = Field(..., min_length=1)
    role: str = "staff"
    active: bool = True
    pin_hash: str = Field(..., min_length=1)
    profile_image_url: str | None = None


class UserPublic(BaseModel):
    user_id: str
    user_name: str
    role: str
    active: bool
    profile_image_url: str | None = None


class ProductUpsert(BaseModel):
    barcode: str = Field(..., min_length=3)
    sku: str | None = None
    name: str
    unit: str = "pcs"
    minimum_stock: int = 0
    category: str | None = None
    location: str | None = None
    opening_balance: int = 0


class UserUpsert(BaseModel):
    requester_id: str = Field(..., min_length=1)
    user_id: str = Field(..., min_length=1)
    user_name: str = Field(..., min_length=1)
    role: str = "staff"
    active: bool = True
    pin: str | None = Field(None, min_length=4)
    profile_image_url: str | None = None


class UserDeleteResponse(BaseModel):
    status: Literal["ok"] = "ok"
    message: str
    deleted_user_id: str
    deleted_movements: int = 0


class LoginRequest(BaseModel):
    user_id: str = Field(..., min_length=1)
    pin: str = Field(..., min_length=4)


class LoginResponse(BaseModel):
    access_token: str
    token_type: Literal["bearer"] = "bearer"
    expires_at: datetime
    user: UserPublic


class ChangePinRequest(BaseModel):
    current_pin: str = Field(..., min_length=4)
    new_pin: str = Field(..., min_length=4)


class ProfileUpdateRequest(BaseModel):
    user_name: str = Field(..., min_length=1)


class StatusMessage(BaseModel):
    status: Literal["ok"] = "ok"
    message: str


class ExportLinkRequest(BaseModel):
    export_name: Literal["products_csv", "users_csv", "movements_csv", "all_xlsx"]
    movement_limit: int = Field(5000, ge=1, le=20000)
    barcode: str | None = None
    actor_id: str | None = None


class ExportLinkResponse(BaseModel):
    url: str
    expires_at: datetime


class ScanRequest(BaseModel):
    barcode: str = Field(..., min_length=3)
    action: MovementType
    quantity: int = Field(..., gt=0)
    actor_id: str = Field(..., min_length=1)
    actor_name: str = Field(..., min_length=1)
    note: str | None = None
    reference: str | None = None
    auto_create_product: bool = False
    product_name: str | None = None
    product_unit: str = "pcs"
    product_minimum_stock: int = 0
    product_category: str | None = None
    product_location: str | None = None
    product_sku: str | None = None


class MovementRecord(BaseModel):
    id: str
    barcode: str
    product_name: str
    action: MovementType
    quantity: int
    before_stock: int
    after_stock: int
    actor_id: str
    actor_name: str
    note: str | None = None
    reference: str | None = None
    created_at: datetime


class OrderItem(BaseModel):
    barcode: str = Field(..., min_length=3)
    product_name: str
    quantity: int = Field(..., gt=0)
    unit: str = "pcs"


class Order(BaseModel):
    id: str
    customer_name: str
    customer_phone: str | None = None
    customer_address: str | None = None
    note: str | None = None
    status: OrderStatus = OrderStatus.NEW
    created_by_id: str
    created_by_name: str
    assigned_to_id: str | None = None
    assigned_to_name: str | None = None
    items: list[OrderItem] = Field(default_factory=list)
    created_at: datetime
    updated_at: datetime


class OrderCreateItem(BaseModel):
    barcode: str = Field(..., min_length=3)
    quantity: int = Field(..., gt=0)


class OrderCreateRequest(BaseModel):
    customer_name: str = Field(..., min_length=1)
    customer_phone: str | None = None
    customer_address: str | None = None
    note: str | None = None
    assigned_to_id: str | None = None
    items: list[OrderCreateItem] = Field(..., min_length=1)


class OrderAssignRequest(BaseModel):
    assigned_to_id: str = Field(..., min_length=1)


class OrderStatusUpdateRequest(BaseModel):
    status: OrderStatus


class SheetMapping(BaseModel):
    spreadsheet_id: str | None = None
    service_account_file: str | None = None
    products_sheet: str = "products"
    movements_sheet: str = "movements"
    users_sheet: str = "users"
    barcode_column: str = "barcode"
    name_column: str = "name"
    stock_column: str = "current_stock"
    minimum_stock_column: str = "minimum_stock"
    sku_column: str = "sku"
    unit_column: str = "unit"
    category_column: str = "category"
    location_column: str = "location"
    user_id_column: str = "user_id"
    user_name_column: str = "user_name"
    user_role_column: str = "role"
    user_active_column: str = "active"
    user_pin_column: str = "pin"
    user_profile_image_column: str = "profile_image_url"


class GoogleSheetBootstrap(BaseModel):
    status: Literal["ready"]
    message: str
    recommended_tabs: list[str]
    expected_product_columns: list[str]
    expected_movement_columns: list[str]
    expected_user_columns: list[str]
    mapping: SheetMapping
    environment_keys: list[str]


class GoogleSheetConfigRequest(BaseModel):
    spreadsheet_id: str
    service_account_file: str | None = None
    products_sheet: str = "products"
    movements_sheet: str = "movements"
    users_sheet: str = "users"
    barcode_column: str = "barcode"
    name_column: str = "name"
    stock_column: str = "current_stock"
    minimum_stock_column: str = "minimum_stock"
    sku_column: str = "sku"
    unit_column: str = "unit"
    category_column: str = "category"
    location_column: str = "location"
    user_id_column: str = "user_id"
    user_name_column: str = "user_name"
    user_role_column: str = "role"
    user_active_column: str = "active"
    user_pin_column: str = "pin"
    user_profile_image_column: str = "profile_image_url"


class GoogleSheetStatus(BaseModel):
    configured: bool
    spreadsheet_id: str | None = None
    service_account_file: str | None = None
    products_sheet: str
    movements_sheet: str
    users_sheet: str
    products_ready: bool
    users_ready: bool
    movement_log_ready: bool
    stock_writeback_ready: bool
    product_autocreate_ready: bool
    message: str


class GoogleSheetSyncResult(BaseModel):
    imported_count: int
    skipped_rows: int
    sheet_name: str
    headers: list[str]
    message: str


class GoogleSheetAppendResult(BaseModel):
    success: bool
    sheet_name: str
    message: str


class GoogleSheetStockUpdateResult(BaseModel):
    success: bool
    sheet_name: str
    barcode: str
    updated_stock: int
    row_number: int
    message: str


class GoogleSheetProductCreateResult(BaseModel):
    success: bool
    sheet_name: str
    barcode: str
    row_number: int | None = None
    message: str


class GoogleSheetBulkSyncResult(BaseModel):
    success: bool
    updated_count: int
    skipped_count: int
    sheet_name: str
    message: str


class ChatAssistantRequest(BaseModel):
    message: str = Field(..., min_length=1, max_length=1000)


class ChatAssistantAction(BaseModel):
    type: Literal["stock_in", "stock_out", "stock_issue"]
    barcode: str
    product_name: str
    quantity: int
    previous_stock: int
    current_stock: int
    low_stock: bool
    movement_id: str


class ChatAssistantResponse(BaseModel):
    message: str
    matched_products: list[Product] = Field(default_factory=list)
    action: ChatAssistantAction | None = None
    download_link: ExportLinkResponse | None = None
    ai_enabled: bool = False
    used_ai: bool = False


class BarcodeSuggestion(BaseModel):
    barcode: str
    format: str
    message: str


class RealtimeEvent(BaseModel):
    type: str
    timestamp: datetime
    payload: dict[str, Any] = Field(default_factory=dict)


app = FastAPI(
    title="Stock Scanner API",
    description=(
        "Backend for a stock-counting mobile app that supports barcode scans, "
        "stock in/out/issue flows, movement history, users, and Google Sheets integration."
    ),
    version="0.3.0",
)

logger = logging.getLogger("stock_scanner_api")
logging.basicConfig(level=os.getenv("LOG_LEVEL", "INFO"))

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


class RealtimeConnectionManager:
    def __init__(self) -> None:
        self._connections: set[WebSocket] = set()

    async def connect(self, websocket: WebSocket) -> None:
        await websocket.accept()
        self._connections.add(websocket)

    def disconnect(self, websocket: WebSocket) -> None:
        self._connections.discard(websocket)

    async def broadcast(self, event: RealtimeEvent) -> None:
        dead_connections: list[WebSocket] = []
        message = event.model_dump(mode="json")
        for websocket in list(self._connections):
            try:
                await websocket.send_json(message)
            except Exception:
                dead_connections.append(websocket)
        for websocket in dead_connections:
            self.disconnect(websocket)


realtime_manager = RealtimeConnectionManager()

UPLOADS_DIR = Path("uploads")
PROFILE_UPLOADS_DIR = UPLOADS_DIR / "profile_images"
PROFILE_UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
app.mount("/uploads", StaticFiles(directory=str(UPLOADS_DIR)), name="uploads")
DB_PATH = Path(os.getenv("STOCK_SCANNER_DB", "stock_scanner.db"))
TOKEN_TTL_HOURS = int(os.getenv("AUTH_TOKEN_TTL_HOURS", "12"))
PIN_HASH_ITERATIONS = int(os.getenv("PIN_HASH_ITERATIONS", "120000"))
EXPORT_LINK_TTL_MINUTES = int(os.getenv("EXPORT_LINK_TTL_MINUTES", "10"))


SEED_PRODUCTS = [
    Product(
        barcode="8850001110012",
        sku="ITM-001",
        name="Printer Paper A4",
        unit="ream",
        minimum_stock=10,
        current_stock=45,
        category="Office",
        location="Rack A1",
    ),
    Product(
        barcode="8850001110013",
        sku="ITM-002",
        name="Hand Sanitizer 500ml",
        unit="bottle",
        minimum_stock=12,
        current_stock=18,
        category="Cleaning",
        location="Rack B2",
    ),
]

SEED_USERS = [
    {
        "user_id": "EMP001",
        "user_name": "Nok",
        "role": "admin",
        "active": True,
        "pin": "1234",
        "profile_image_url": None,
    },
    {
        "user_id": "EMP002",
        "user_name": "Mek",
        "role": "staff",
        "active": True,
        "pin": "1234",
        "profile_image_url": None,
    },
]

products: dict[str, Product] = {}
users: dict[str, User] = {}

movements: list[MovementRecord] = []
orders: list[Order] = []
sheet_mapping = SheetMapping(
    spreadsheet_id=os.getenv("GOOGLE_SHEETS_SPREADSHEET_ID"),
    service_account_file=os.getenv("GOOGLE_SERVICE_ACCOUNT_FILE"),
)


def utc_now() -> datetime:
    return datetime.now(timezone.utc)


def db_connection() -> sqlite3.Connection:
    if DB_PATH.parent != Path("."):
        DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    connection = sqlite3.connect(DB_PATH)
    connection.row_factory = sqlite3.Row
    return connection


def hash_pin(pin: str) -> str:
    salt = secrets.token_hex(16)
    derived = hashlib.pbkdf2_hmac(
        "sha256",
        pin.encode("utf-8"),
        salt.encode("utf-8"),
        PIN_HASH_ITERATIONS,
    ).hex()
    return f"pbkdf2_sha256${PIN_HASH_ITERATIONS}${salt}${derived}"


def verify_pin(pin: str, stored_hash: str) -> bool:
    try:
        algorithm, iterations_text, salt, expected_hash = stored_hash.split("$", 3)
    except ValueError:
        return False
    if algorithm != "pbkdf2_sha256":
        return False
    iterations = int(iterations_text)
    candidate = hashlib.pbkdf2_hmac(
        "sha256",
        pin.encode("utf-8"),
        salt.encode("utf-8"),
        iterations,
    ).hex()
    return hmac.compare_digest(candidate, expected_hash)


def normalize_datetime(value: str | datetime) -> datetime:
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=timezone.utc)
    normalized = value.replace("Z", "+00:00")
    parsed = datetime.fromisoformat(normalized)
    return parsed if parsed.tzinfo else parsed.replace(tzinfo=timezone.utc)


def normalize_header(value: str) -> str:
    return value.strip().lower().replace(" ", "_")


def parse_int(value: Any, default: int = 0) -> int:
    if value is None or value == "":
        return default
    if isinstance(value, (int, float)):
        return int(value)
    try:
        return int(float(str(value).strip()))
    except ValueError:
        return default


def parse_bool(value: Any, default: bool = True) -> bool:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"true", "1", "yes", "y", "active"}


def parse_allowed_origins(value: str | None) -> list[str]:
    if not value or not value.strip():
        return ["*"]
    origins = [item.strip() for item in value.split(",") if item.strip()]
    return origins or ["*"]


def generate_next_internal_barcode() -> str:
    prefix = "STK"
    max_number = 0
    for barcode in products:
        normalized = barcode.strip().upper()
        if normalized.startswith(prefix) and normalized[len(prefix) :].isdigit():
            max_number = max(max_number, int(normalized[len(prefix) :]))
    return f"{prefix}{max_number + 1:06d}"


def to_public_user(user: User) -> UserPublic:
    return UserPublic(
        user_id=user.user_id,
        user_name=user.user_name,
        role=user.role,
        active=user.active,
        profile_image_url=user.profile_image_url,
    )


def init_database() -> None:
    with db_connection() as connection:
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS products (
                barcode TEXT PRIMARY KEY,
                sku TEXT,
                name TEXT NOT NULL,
                unit TEXT NOT NULL,
                minimum_stock INTEGER NOT NULL DEFAULT 0,
                current_stock INTEGER NOT NULL DEFAULT 0,
                category TEXT,
                location TEXT
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS users (
                user_id TEXT PRIMARY KEY,
                user_name TEXT NOT NULL,
                role TEXT NOT NULL DEFAULT 'staff',
                active INTEGER NOT NULL DEFAULT 1,
                pin_hash TEXT NOT NULL,
                profile_image_url TEXT
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS movements (
                id TEXT PRIMARY KEY,
                barcode TEXT NOT NULL,
                product_name TEXT NOT NULL,
                action TEXT NOT NULL,
                quantity INTEGER NOT NULL,
                before_stock INTEGER NOT NULL,
                after_stock INTEGER NOT NULL,
                actor_id TEXT NOT NULL,
                actor_name TEXT NOT NULL,
                note TEXT,
                reference TEXT,
                created_at TEXT NOT NULL
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS auth_sessions (
                token TEXT PRIMARY KEY,
                user_id TEXT NOT NULL,
                created_at TEXT NOT NULL,
                expires_at TEXT NOT NULL
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS export_tokens (
                token TEXT PRIMARY KEY,
                user_id TEXT NOT NULL,
                export_name TEXT NOT NULL,
                params_json TEXT NOT NULL,
                created_at TEXT NOT NULL,
                expires_at TEXT NOT NULL
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS orders (
                id TEXT PRIMARY KEY,
                customer_name TEXT NOT NULL,
                customer_phone TEXT,
                customer_address TEXT,
                note TEXT,
                status TEXT NOT NULL,
                created_by_id TEXT NOT NULL,
                created_by_name TEXT NOT NULL,
                assigned_to_id TEXT,
                assigned_to_name TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS order_items (
                id TEXT PRIMARY KEY,
                order_id TEXT NOT NULL,
                barcode TEXT NOT NULL,
                product_name TEXT NOT NULL,
                quantity INTEGER NOT NULL,
                unit TEXT NOT NULL
            )
            """
        )

        existing_product_count = connection.execute(
            "SELECT COUNT(*) AS count FROM products"
        ).fetchone()["count"]
        if existing_product_count == 0:
            connection.executemany(
                """
                INSERT INTO products (
                    barcode, sku, name, unit, minimum_stock, current_stock, category, location
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                [
                    (
                        product.barcode,
                        product.sku,
                        product.name,
                        product.unit,
                        product.minimum_stock,
                        product.current_stock,
                        product.category,
                        product.location,
                    )
                    for product in SEED_PRODUCTS
                ],
            )

        existing_user_count = connection.execute(
            "SELECT COUNT(*) AS count FROM users"
        ).fetchone()["count"]
        if existing_user_count == 0:
            connection.executemany(
                """
                INSERT INTO users (
                    user_id, user_name, role, active, pin_hash, profile_image_url
                ) VALUES (?, ?, ?, ?, ?, ?)
                """,
                [
                    (
                        seed_user["user_id"],
                        seed_user["user_name"],
                        seed_user["role"],
                        1 if seed_user["active"] else 0,
                        hash_pin(str(seed_user["pin"])),
                        seed_user["profile_image_url"],
                    )
                    for seed_user in SEED_USERS
                ],
            )


def load_state_from_db() -> None:
    global products, users, movements, orders
    with db_connection() as connection:
        product_rows = connection.execute(
            """
            SELECT barcode, sku, name, unit, minimum_stock, current_stock, category, location
            FROM products
            ORDER BY name COLLATE NOCASE
            """
        ).fetchall()
        user_rows = connection.execute(
            """
            SELECT user_id, user_name, role, active, pin_hash, profile_image_url
            FROM users
            ORDER BY user_name COLLATE NOCASE
            """
        ).fetchall()
        movement_rows = connection.execute(
            """
            SELECT id, barcode, product_name, action, quantity, before_stock, after_stock,
                   actor_id, actor_name, note, reference, created_at
            FROM movements
            ORDER BY created_at DESC
            """
        ).fetchall()
        order_rows = connection.execute(
            """
            SELECT id, customer_name, customer_phone, customer_address, note, status,
                   created_by_id, created_by_name, assigned_to_id, assigned_to_name,
                   created_at, updated_at
            FROM orders
            ORDER BY created_at DESC
            """
        ).fetchall()
        order_item_rows = connection.execute(
            """
            SELECT order_id, barcode, product_name, quantity, unit
            FROM order_items
            ORDER BY rowid ASC
            """
        ).fetchall()

    products = {
        row["barcode"]: Product(
            barcode=row["barcode"],
            sku=row["sku"],
            name=row["name"],
            unit=row["unit"],
            minimum_stock=int(row["minimum_stock"]),
            current_stock=int(row["current_stock"]),
            category=row["category"],
            location=row["location"],
        )
        for row in product_rows
    }
    users = {
        row["user_id"]: User(
            user_id=row["user_id"],
            user_name=row["user_name"],
            role=row["role"],
            active=bool(row["active"]),
            pin_hash=row["pin_hash"],
            profile_image_url=row["profile_image_url"],
        )
        for row in user_rows
    }
    movements = [
        MovementRecord(
            id=row["id"],
            barcode=row["barcode"],
            product_name=row["product_name"],
            action=MovementType(row["action"]),
            quantity=int(row["quantity"]),
            before_stock=int(row["before_stock"]),
            after_stock=int(row["after_stock"]),
            actor_id=row["actor_id"],
            actor_name=row["actor_name"],
            note=row["note"],
            reference=row["reference"],
            created_at=normalize_datetime(row["created_at"]),
        )
        for row in movement_rows
    ]
    items_by_order: dict[str, list[OrderItem]] = {}
    for row in order_item_rows:
        items_by_order.setdefault(row["order_id"], []).append(
            OrderItem(
                barcode=row["barcode"],
                product_name=row["product_name"],
                quantity=int(row["quantity"]),
                unit=row["unit"],
            )
        )
    orders = [
        Order(
            id=row["id"],
            customer_name=row["customer_name"],
            customer_phone=row["customer_phone"],
            customer_address=row["customer_address"],
            note=row["note"],
            status=OrderStatus(row["status"]),
            created_by_id=row["created_by_id"],
            created_by_name=row["created_by_name"],
            assigned_to_id=row["assigned_to_id"],
            assigned_to_name=row["assigned_to_name"],
            items=items_by_order.get(row["id"], []),
            created_at=normalize_datetime(row["created_at"]),
            updated_at=normalize_datetime(row["updated_at"]),
        )
        for row in order_rows
    ]


def save_product(product: Product) -> None:
    with db_connection() as connection:
        connection.execute(
            """
            INSERT INTO products (
                barcode, sku, name, unit, minimum_stock, current_stock, category, location
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(barcode) DO UPDATE SET
                sku = excluded.sku,
                name = excluded.name,
                unit = excluded.unit,
                minimum_stock = excluded.minimum_stock,
                current_stock = excluded.current_stock,
                category = excluded.category,
                location = excluded.location
            """,
            (
                product.barcode,
                product.sku,
                product.name,
                product.unit,
                product.minimum_stock,
                product.current_stock,
                product.category,
                product.location,
            ),
        )


def save_user(user: User) -> None:
    with db_connection() as connection:
        connection.execute(
            """
            INSERT INTO users (
                user_id, user_name, role, active, pin_hash, profile_image_url
            ) VALUES (?, ?, ?, ?, ?, ?)
            ON CONFLICT(user_id) DO UPDATE SET
                user_name = excluded.user_name,
                role = excluded.role,
                active = excluded.active,
                pin_hash = excluded.pin_hash,
                profile_image_url = excluded.profile_image_url
            """,
            (
                user.user_id,
                user.user_name,
                user.role,
                1 if user.active else 0,
                user.pin_hash,
                user.profile_image_url,
            ),
        )


def active_admin_count(exclude_user_id: str | None = None) -> int:
    return sum(
        1
        for user in users.values()
        if user.active
        and user.role.strip().lower() == "admin"
        and user.user_id != exclude_user_id
    )


def ensure_admin_account_retained(
    target_user_id: str,
    target_role: str,
    target_active: bool,
) -> None:
    existing = users.get(target_user_id)
    if not existing:
        return
    if existing.role.strip().lower() != "admin" or not existing.active:
        return
    if target_role.strip().lower() == "admin" and target_active:
        return
    if active_admin_count(exclude_user_id=target_user_id) == 0:
        raise HTTPException(
            status_code=400,
            detail="The system must keep at least one active admin account.",
        )


def save_movement(record: MovementRecord) -> None:
    with db_connection() as connection:
        connection.execute(
            """
            INSERT INTO movements (
                id, barcode, product_name, action, quantity, before_stock, after_stock,
                actor_id, actor_name, note, reference, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                record.id,
                record.barcode,
                record.product_name,
                record.action.value,
                record.quantity,
                record.before_stock,
                record.after_stock,
                record.actor_id,
                record.actor_name,
                record.note,
                record.reference,
                record.created_at.isoformat(),
            ),
        )


def save_order(order: Order) -> None:
    with db_connection() as connection:
        connection.execute(
            """
            INSERT INTO orders (
                id, customer_name, customer_phone, customer_address, note, status,
                created_by_id, created_by_name, assigned_to_id, assigned_to_name,
                created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(id) DO UPDATE SET
                customer_name = excluded.customer_name,
                customer_phone = excluded.customer_phone,
                customer_address = excluded.customer_address,
                note = excluded.note,
                status = excluded.status,
                created_by_id = excluded.created_by_id,
                created_by_name = excluded.created_by_name,
                assigned_to_id = excluded.assigned_to_id,
                assigned_to_name = excluded.assigned_to_name,
                created_at = excluded.created_at,
                updated_at = excluded.updated_at
            """,
            (
                order.id,
                order.customer_name,
                order.customer_phone,
                order.customer_address,
                order.note,
                order.status.value,
                order.created_by_id,
                order.created_by_name,
                order.assigned_to_id,
                order.assigned_to_name,
                order.created_at.isoformat(),
                order.updated_at.isoformat(),
            ),
        )
        connection.execute("DELETE FROM order_items WHERE order_id = ?", (order.id,))
        connection.executemany(
            """
            INSERT INTO order_items (id, order_id, barcode, product_name, quantity, unit)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    str(uuid4()),
                    order.id,
                    item.barcode,
                    item.product_name,
                    item.quantity,
                    item.unit,
                )
                for item in order.items
            ],
        )


def create_session(user_id: str) -> tuple[str, datetime]:
    token = secrets.token_urlsafe(32)
    created_at = utc_now()
    expires_at = created_at + timedelta(hours=TOKEN_TTL_HOURS)
    with db_connection() as connection:
        connection.execute(
            "DELETE FROM auth_sessions WHERE expires_at <= ?",
            (created_at.isoformat(),),
        )
        connection.execute(
            """
            INSERT INTO auth_sessions (token, user_id, created_at, expires_at)
            VALUES (?, ?, ?, ?)
            """,
            (token, user_id, created_at.isoformat(), expires_at.isoformat()),
        )
    return token, expires_at


def user_from_token(token: str | None) -> User | None:
    if not token:
        return None
    with db_connection() as connection:
        row = connection.execute(
            """
            SELECT s.user_id, s.expires_at
            FROM auth_sessions AS s
            WHERE s.token = ?
            """,
            (token,),
        ).fetchone()
        if not row:
            return None
        if normalize_datetime(row["expires_at"]) <= utc_now():
            connection.execute("DELETE FROM auth_sessions WHERE token = ?", (token,))
            return None
    return users.get(row["user_id"])


def delete_session(token: str) -> None:
    with db_connection() as connection:
        connection.execute("DELETE FROM auth_sessions WHERE token = ?", (token,))


def delete_user_record(user_id: str, delete_movements: bool = False) -> int:
    removed_movements = 0
    with db_connection() as connection:
        connection.execute("DELETE FROM auth_sessions WHERE user_id = ?", (user_id,))
        connection.execute("DELETE FROM export_tokens WHERE user_id = ?", (user_id,))
        if delete_movements:
            cursor = connection.execute("DELETE FROM movements WHERE actor_id = ?", (user_id,))
            removed_movements = cursor.rowcount if cursor.rowcount != -1 else 0
        connection.execute("DELETE FROM users WHERE user_id = ?", (user_id,))
    return removed_movements


def create_export_token(user_id: str, export_name: str, params: dict[str, Any]) -> tuple[str, datetime]:
    token = secrets.token_urlsafe(32)
    created_at = utc_now()
    expires_at = created_at + timedelta(minutes=EXPORT_LINK_TTL_MINUTES)
    with db_connection() as connection:
        connection.execute(
            "DELETE FROM export_tokens WHERE expires_at <= ?",
            (created_at.isoformat(),),
        )
        connection.execute(
            """
            INSERT INTO export_tokens (token, user_id, export_name, params_json, created_at, expires_at)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                token,
                user_id,
                export_name,
                json.dumps(params),
                created_at.isoformat(),
                expires_at.isoformat(),
            ),
        )
    return token, expires_at


def consume_export_token(token: str) -> tuple[str, dict[str, Any], User]:
    with db_connection() as connection:
        row = connection.execute(
            """
            SELECT token, user_id, export_name, params_json, expires_at
            FROM export_tokens
            WHERE token = ?
            """,
            (token,),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Export link not found or already used.")
        if normalize_datetime(row["expires_at"]) <= utc_now():
            connection.execute("DELETE FROM export_tokens WHERE token = ?", (token,))
            raise HTTPException(status_code=410, detail="Export link has expired.")
        connection.execute("DELETE FROM export_tokens WHERE token = ?", (token,))

    user = get_user_or_404(row["user_id"])
    if user.role.strip().lower() != "admin":
        raise HTTPException(status_code=403, detail="Admin permission required.")
    return row["export_name"], json.loads(row["params_json"]), user


def bearer_token_from_request(request: Request) -> str | None:
    authorization = request.headers.get("Authorization", "").strip()
    if not authorization:
        return None
    scheme, _, token = authorization.partition(" ")
    if scheme.lower() != "bearer" or not token:
        raise HTTPException(status_code=401, detail="Invalid authorization header.")
    return token.strip()


def resolve_request_user(request: Request, requester_id: str | None = None) -> User:
    token = bearer_token_from_request(request)
    if token:
        user = user_from_token(token)
        if not user:
            raise HTTPException(status_code=401, detail="Invalid or expired access token.")
        if not user.active:
            raise HTTPException(status_code=403, detail=f"User {user.user_id} is inactive.")
        return user
    if requester_id:
        return get_user_or_404(requester_id)
    raise HTTPException(status_code=401, detail="Authentication required.")


def require_admin_request(request: Request, requester_id: str | None = None) -> User:
    user = resolve_request_user(request, requester_id)
    if user.role.strip().lower() != "admin":
        raise HTTPException(status_code=403, detail="Admin permission required.")
    return user


class GoogleSheetsClient:
    def __init__(self, mapping: SheetMapping):
        self.mapping = mapping

    def _validate_config(self) -> tuple[str, str]:
        spreadsheet_id = self.mapping.spreadsheet_id or os.getenv("GOOGLE_SHEETS_SPREADSHEET_ID")
        service_account_file = self.mapping.service_account_file or os.getenv(
            "GOOGLE_SERVICE_ACCOUNT_FILE"
        )
        if not spreadsheet_id:
            raise HTTPException(
                status_code=400,
                detail="Google Sheets is not configured. Missing spreadsheet_id.",
            )
        if not service_account_file:
            raise HTTPException(
                status_code=400,
                detail="Google Sheets is not configured. Missing service_account_file.",
            )
        if not os.path.exists(service_account_file):
            raise HTTPException(
                status_code=400,
                detail=f"Service account file not found: {service_account_file}",
            )
        return spreadsheet_id, service_account_file

    def _service(self) -> tuple[Any, str]:
        spreadsheet_id, service_account_file = self._validate_config()
        try:
            from google.oauth2.service_account import Credentials
            from googleapiclient.discovery import build
        except ImportError as exc:
            raise HTTPException(
                status_code=500,
                detail=(
                    "Missing Google Sheets dependencies. Install "
                    "'google-api-python-client' and 'google-auth'."
                ),
            ) from exc

        scopes = ["https://www.googleapis.com/auth/spreadsheets"]
        credentials = Credentials.from_service_account_file(
            service_account_file,
            scopes=scopes,
        )
        service = build("sheets", "v4", credentials=credentials, cache_discovery=False)
        return service, spreadsheet_id

    def _get_sheet_values(self, service: Any, spreadsheet_id: str, sheet_name: str) -> list[list[Any]]:
        response = (
            service.spreadsheets()
            .values()
            .get(
                spreadsheetId=spreadsheet_id,
                range=f"{sheet_name}!A1:ZZ",
            )
            .execute()
        )
        return response.get("values", [])

    def _column_letter(self, column_number: int) -> str:
        result = ""
        current = column_number
        while current > 0:
            current, remainder = divmod(current - 1, 26)
            result = chr(65 + remainder) + result
        return result

    def _find_product_row(self, values: list[list[Any]], barcode: str) -> tuple[int, int]:
        if not values:
            raise HTTPException(status_code=404, detail="Products sheet is empty.")

        headers = [normalize_header(item) for item in values[0]]
        try:
            barcode_column_index = headers.index(self.mapping.barcode_column)
            stock_column_index = headers.index(self.mapping.stock_column)
        except ValueError as exc:
            raise HTTPException(
                status_code=400,
                detail=(
                    "Products sheet headers do not match the configured barcode_column "
                    "or stock_column."
                ),
            ) from exc

        for row_index, raw_row in enumerate(values[1:], start=2):
            row_barcode = ""
            if barcode_column_index < len(raw_row):
                row_barcode = str(raw_row[barcode_column_index]).strip()
            if row_barcode == barcode:
                return row_index, stock_column_index + 1

        raise HTTPException(
            status_code=404,
            detail=f"Barcode {barcode} was not found in the products sheet.",
        )

    def read_products(self) -> GoogleSheetSyncResult:
        service, spreadsheet_id = self._service()
        values = self._get_sheet_values(service, spreadsheet_id, self.mapping.products_sheet)
        if not values:
            raise HTTPException(status_code=404, detail="Products sheet is empty.")

        headers = [normalize_header(item) for item in values[0]]
        imported_count = 0
        skipped_rows = 0

        for raw_row in values[1:]:
            row = dict(zip(headers, raw_row))
            barcode = str(row.get(self.mapping.barcode_column, "")).strip()
            name = str(row.get(self.mapping.name_column, "")).strip()
            if not barcode or not name:
                skipped_rows += 1
                continue

            products[barcode] = Product(
                barcode=barcode,
                sku=str(row.get(self.mapping.sku_column, "")).strip() or None,
                name=name,
                unit=str(row.get(self.mapping.unit_column, "")).strip() or "pcs",
                minimum_stock=parse_int(row.get(self.mapping.minimum_stock_column)),
                current_stock=parse_int(row.get(self.mapping.stock_column)),
                category=str(row.get(self.mapping.category_column, "")).strip() or None,
                location=str(row.get(self.mapping.location_column, "")).strip() or None,
            )
            save_product(products[barcode])
            imported_count += 1

        return GoogleSheetSyncResult(
            imported_count=imported_count,
            skipped_rows=skipped_rows,
            sheet_name=self.mapping.products_sheet,
            headers=headers,
            message="Products imported from Google Sheets.",
        )

    def read_users(self) -> GoogleSheetSyncResult:
        service, spreadsheet_id = self._service()
        values = self._get_sheet_values(service, spreadsheet_id, self.mapping.users_sheet)
        if not values:
            raise HTTPException(status_code=404, detail="Users sheet is empty.")

        headers = [normalize_header(item) for item in values[0]]
        imported_count = 0
        skipped_rows = 0

        for raw_row in values[1:]:
            row = dict(zip(headers, raw_row))
            user_id = str(row.get(self.mapping.user_id_column, "")).strip()
            user_name = str(row.get(self.mapping.user_name_column, "")).strip()
            if not user_id or not user_name:
                skipped_rows += 1
                continue

            users[user_id] = User(
                user_id=user_id,
                user_name=user_name,
                role=str(row.get(self.mapping.user_role_column, "")).strip() or "staff",
                active=parse_bool(row.get(self.mapping.user_active_column), default=True),
                pin_hash=hash_pin(
                    str(row.get(self.mapping.user_pin_column, "")).strip() or "1234"
                ),
                profile_image_url=(
                    str(row.get(self.mapping.user_profile_image_column, "")).strip() or None
                ),
            )
            save_user(users[user_id])
            imported_count += 1

        return GoogleSheetSyncResult(
            imported_count=imported_count,
            skipped_rows=skipped_rows,
            sheet_name=self.mapping.users_sheet,
            headers=headers,
            message="Users imported from Google Sheets.",
        )

    def append_product(self, product: Product) -> GoogleSheetProductCreateResult:
        service, spreadsheet_id = self._service()
        values = [
            [
                product.barcode,
                product.sku or "",
                product.name,
                product.unit,
                product.current_stock,
                product.minimum_stock,
                product.category or "",
                product.location or "",
            ]
        ]
        (
            service.spreadsheets()
            .values()
            .append(
                spreadsheetId=spreadsheet_id,
                range=f"{self.mapping.products_sheet}!A1",
                valueInputOption="USER_ENTERED",
                insertDataOption="INSERT_ROWS",
                body={"values": values},
            )
            .execute()
        )
        values_after = self._get_sheet_values(service, spreadsheet_id, self.mapping.products_sheet)
        row_number = len(values_after)
        return GoogleSheetProductCreateResult(
            success=True,
            sheet_name=self.mapping.products_sheet,
            barcode=product.barcode,
            row_number=row_number,
            message="Product appended to Google Sheets.",
        )

    def upsert_user(self, user: User) -> GoogleSheetAppendResult:
        service, spreadsheet_id = self._service()
        existing_values = self._get_sheet_values(service, spreadsheet_id, self.mapping.users_sheet)
        if not existing_values:
            raise HTTPException(status_code=404, detail="Users sheet is empty.")

        headers = [normalize_header(item) for item in existing_values[0]]
        try:
            user_id_column_index = headers.index(self.mapping.user_id_column)
        except ValueError as exc:
            raise HTTPException(
                status_code=400,
                detail="Users sheet headers do not match the configured user_id_column.",
            ) from exc

        header_value_map = {
            self.mapping.user_id_column: user.user_id,
            self.mapping.user_name_column: user.user_name,
            self.mapping.user_role_column: user.role,
            self.mapping.user_active_column: "TRUE" if user.active else "FALSE",
            self.mapping.user_pin_column: user.pin_hash,
            self.mapping.user_profile_image_column: user.profile_image_url or "",
        }
        ordered_values = [[header_value_map.get(header, "") for header in headers]]

        existing_row_number: int | None = None
        for row_index, raw_row in enumerate(existing_values[1:], start=2):
            row_user_id = ""
            if user_id_column_index < len(raw_row):
                row_user_id = str(raw_row[user_id_column_index]).strip()
            if row_user_id == user.user_id:
                existing_row_number = row_index
                break

        if existing_row_number is None:
            (
                service.spreadsheets()
                .values()
                .append(
                    spreadsheetId=spreadsheet_id,
                    range=f"{self.mapping.users_sheet}!A1",
                    valueInputOption="USER_ENTERED",
                    insertDataOption="INSERT_ROWS",
                    body={"values": ordered_values},
                )
                .execute()
            )
        else:
            last_column_letter = self._column_letter(len(headers))
            (
                service.spreadsheets()
                .values()
                .update(
                    spreadsheetId=spreadsheet_id,
                    range=(
                        f"{self.mapping.users_sheet}!A{existing_row_number}:"
                        f"{last_column_letter}{existing_row_number}"
                    ),
                    valueInputOption="USER_ENTERED",
                    body={"values": ordered_values},
                )
                .execute()
            )

        return GoogleSheetAppendResult(
            success=True,
            sheet_name=self.mapping.users_sheet,
            message="User written to Google Sheets.",
        )

    def update_product_stock(self, product: Product) -> GoogleSheetStockUpdateResult:
        service, spreadsheet_id = self._service()
        values = self._get_sheet_values(service, spreadsheet_id, self.mapping.products_sheet)
        row_number, stock_column_number = self._find_product_row(values, product.barcode)
        stock_column_letter = self._column_letter(stock_column_number)

        (
            service.spreadsheets()
            .values()
            .update(
                spreadsheetId=spreadsheet_id,
                range=f"{self.mapping.products_sheet}!{stock_column_letter}{row_number}",
                valueInputOption="USER_ENTERED",
                body={"values": [[product.current_stock]]},
            )
            .execute()
        )

        return GoogleSheetStockUpdateResult(
            success=True,
            sheet_name=self.mapping.products_sheet,
            barcode=product.barcode,
            updated_stock=product.current_stock,
            row_number=row_number,
            message="Product stock updated in Google Sheets.",
        )

    def sync_all_product_stocks(self) -> GoogleSheetBulkSyncResult:
        updated_count = 0
        skipped_count = 0

        for product in products.values():
            try:
                self.update_product_stock(product)
                updated_count += 1
            except HTTPException:
                skipped_count += 1

        return GoogleSheetBulkSyncResult(
            success=True,
            updated_count=updated_count,
            skipped_count=skipped_count,
            sheet_name=self.mapping.products_sheet,
            message="Local stock balances were pushed to Google Sheets.",
        )

    def append_movement(self, record: MovementRecord) -> GoogleSheetAppendResult:
        service, spreadsheet_id = self._service()
        values = [
            [
                record.id,
                record.created_at.isoformat(),
                record.barcode,
                record.product_name,
                record.action.value,
                record.quantity,
                record.before_stock,
                record.after_stock,
                record.actor_id,
                record.actor_name,
                record.note or "",
                record.reference or "",
            ]
        ]
        (
            service.spreadsheets()
            .values()
            .append(
                spreadsheetId=spreadsheet_id,
                range=f"{self.mapping.movements_sheet}!A1",
                valueInputOption="USER_ENTERED",
                insertDataOption="INSERT_ROWS",
                body={"values": values},
            )
            .execute()
        )
        return GoogleSheetAppendResult(
            success=True,
            sheet_name=self.mapping.movements_sheet,
            message="Movement appended to Google Sheets.",
        )

    def status(self) -> GoogleSheetStatus:
        configured = bool(
            (self.mapping.spreadsheet_id or os.getenv("GOOGLE_SHEETS_SPREADSHEET_ID"))
            and (self.mapping.service_account_file or os.getenv("GOOGLE_SERVICE_ACCOUNT_FILE"))
        )
        return GoogleSheetStatus(
            configured=configured,
            spreadsheet_id=self.mapping.spreadsheet_id,
            service_account_file=self.mapping.service_account_file,
            products_sheet=self.mapping.products_sheet,
            movements_sheet=self.mapping.movements_sheet,
            users_sheet=self.mapping.users_sheet,
            products_ready=configured,
            users_ready=configured,
            movement_log_ready=configured,
            stock_writeback_ready=configured,
            product_autocreate_ready=configured,
            message=(
                "Ready to sync with Google Sheets."
                if configured
                else "Waiting for spreadsheet_id and service_account_file."
            ),
        )


def create_notification(record: MovementRecord) -> dict:
    return {
        "title": f"Stock {record.action.value}",
        "message": (
            f"{record.actor_name} performed '{record.action.value}' on "
            f"{record.product_name} x{record.quantity}"
        ),
        "movement_id": record.id,
        "barcode": record.barcode,
        "created_at": record.created_at,
    }


async def broadcast_realtime_event(event_type: str, payload: dict[str, Any]) -> None:
    await realtime_manager.broadcast(
        RealtimeEvent(
            type=event_type,
            timestamp=utc_now(),
            payload=payload,
        )
    )


def sheet_client() -> GoogleSheetsClient:
    return GoogleSheetsClient(sheet_mapping)


def get_user_or_404(user_id: str) -> User:
    user = users.get(user_id)
    if not user:
        raise HTTPException(status_code=404, detail=f"User {user_id} not found.")
    if not user.active:
        raise HTTPException(status_code=403, detail=f"User {user_id} is inactive.")
    return user


def require_admin(user_id: str) -> User:
    user = get_user_or_404(user_id)
    if user.role.strip().lower() != "admin":
        raise HTTPException(status_code=403, detail="Admin permission required.")
    return user


def can_manage_user_profile(requester_id: str, target_user_id: str) -> User:
    requester = get_user_or_404(requester_id)
    if requester.user_id == target_user_id:
        return requester
    if requester.role.strip().lower() == "admin":
        return requester
    raise HTTPException(status_code=403, detail="You can update only your own profile image.")


def build_profile_image_url(filename: str) -> str:
    return f"/uploads/profile_images/{filename}"


def get_order_or_404(order_id: str) -> Order:
    for order in orders:
        if order.id == order_id:
            return order
    raise HTTPException(status_code=404, detail=f"Order {order_id} not found.")


def ai_chat_enabled() -> bool:
    return bool(os.getenv("OPENAI_API_KEY", "").strip())


def normalize_chat_text(value: str) -> str:
    return "".join(ch for ch in value.strip().lower() if ch.isalnum() or "\u0e00" <= ch <= "\u0e7f")


def strip_chat_noise(value: str) -> str:
    normalized = value
    for token in [
        "มี",
        "สินค้า",
        "ตัว",
        "นี้",
        "อัน",
        "รายการ",
        "ตอนนี้",
        "หน่อย",
        "บ้าง",
        "ไหม",
        "หรือยัง",
        "หรือเปล่า",
        "เท่าไหร่",
        "เท่าไร",
        "กี่ชิ้น",
        "กี่อัน",
        "กี่หน่วย",
        "คงเหลือ",
        "เหลือ",
        "สต๊อก",
        "สตอก",
    ]:
        normalized = normalized.replace(token, "")
    return normalized.strip()


def summarize_product(product: Product) -> str:
    status = "LOW" if product.current_stock <= product.minimum_stock else "OK"
    return (
        f"{product.name} | barcode={product.barcode} | sku={product.sku or '-'} | "
        f"stock={product.current_stock} {product.unit} | min={product.minimum_stock} | "
        f"category={product.category or '-'} | location={product.location or '-'} | status={status}"
    )


def find_chat_products(question: str) -> list[Product]:
    normalized_question = normalize_chat_text(question)
    simplified_question = normalize_chat_text(strip_chat_noise(question))
    compact_question = question.replace(" ", "").strip()
    matches: list[Product] = []

    for product in products.values():
        fields = [
            product.name,
            product.barcode,
            product.sku or "",
            product.category or "",
            product.location or "",
        ]
        normalized_fields = [normalize_chat_text(field) for field in fields]
        if (
            compact_question == product.barcode
            or any(field and normalized_question in field for field in normalized_fields)
            or any(field and field in normalized_question for field in normalized_fields)
            or any(field and simplified_question and simplified_question in field for field in normalized_fields)
        ):
            matches.append(product)

    matches.sort(
        key=lambda item: (
            normalize_chat_text(item.name) != normalized_question,
            item.name.lower(),
        )
    )
    return matches


def build_stock_summary_payload() -> dict[str, Any]:
    low_stock_items = [item for item in products.values() if item.current_stock <= item.minimum_stock]
    return {
        "total_products": len(products),
        "total_units": sum(item.current_stock for item in products.values()),
        "low_stock_count": len(low_stock_items),
        "low_stock_items": low_stock_items,
    }


def detect_chat_action(message: str) -> tuple[MovementType, int, str] | None:
    lowered = message.strip().lower()
    intent_map = {
        MovementType.IN: ["เพิ่ม", "รับเข้า", "เติม", "stockin", "นำเข้า", "เอาเข้า", "เพิ่มสต๊อก", "เพิ่มสตอก"],
        MovementType.OUT: ["เบิก", "ตัด", "ลด", "จ่ายออก", "stockout", "เอาออก", "ลดสต๊อก", "ลดสตอก", "ตัดสต๊อก", "ตัดสตอก"],
        MovementType.ISSUE: ["issue", "ใช้ไป", "นำออกใช้", "หยิบใช้", "เบิกใช้"],
    }
    selected_action: MovementType | None = None
    for action, keywords in intent_map.items():
        if any(keyword in lowered for keyword in keywords):
            selected_action = action
            break
    if selected_action is None:
        return None

    quantity = None
    for token in message.replace(",", " ").split():
        if token.isdigit():
            quantity = int(token)
            break
    if quantity is None or quantity <= 0:
        return None

    product_hint = message
    for keyword_list in intent_map.values():
        for keyword in keyword_list:
            product_hint = product_hint.replace(keyword, " ")
            product_hint = product_hint.replace(keyword.capitalize(), " ")
    for token in product_hint.split():
        if token.isdigit():
            product_hint = product_hint.replace(token, " ")
    cleaned_hint = product_hint.strip()
    if not cleaned_hint:
        return None
    return selected_action, quantity, cleaned_hint


def execute_chat_stock_action(message: str, user: User) -> tuple[str, list[Product], ChatAssistantAction | None, dict | None]:
    detected = detect_chat_action(message)
    if not detected:
        return "", [], None, None

    action, quantity, product_hint = detected
    matches = find_chat_products(product_hint)
    if not matches:
        return (
            "ยังหาสินค้าที่ต้องการสั่งงานไม่เจอ ลองระบุชื่อหรือบาร์โค้ดให้ชัดขึ้น",
            [],
            None,
            None,
        )
    if len(matches) > 1:
        preview = ", ".join(item.name for item in matches[:5])
        return (
            f"เจอหลายรายการที่ใกล้เคียงกัน: {preview} ลองระบุให้ชัดขึ้นก่อนสั่งงาน",
            matches[:5],
            None,
            None,
        )

    product = matches[0]
    before_stock = product.current_stock
    delta = quantity if action == MovementType.IN else -quantity
    after_stock = before_stock + delta
    if after_stock < 0:
        return (
            f"สั่งงานไม่ได้ เพราะ {product.name} คงเหลือ {before_stock} {product.unit} แต่ขอใช้ {quantity} {product.unit}",
            [product],
            None,
            None,
        )

    product.current_stock = after_stock
    products[product.barcode] = product
    save_product(product)

    record = MovementRecord(
        id=str(uuid4()),
        barcode=product.barcode,
        product_name=product.name,
        action=action,
        quantity=quantity,
        before_stock=before_stock,
        after_stock=after_stock,
        actor_id=user.user_id,
        actor_name=user.user_name,
        note="Chat assistant action",
        reference=message.strip(),
        created_at=utc_now(),
    )
    movements.insert(0, record)
    save_movement(record)
    try:
        client = sheet_client()
        if client.status().configured:
            client.append_movement(record)
            client.update_product_stock(product)
    except Exception as exc:
        logger.warning("Failed to sync chat action for %s: %s", product.barcode, exc)
    notification = create_notification(record)
    return (
        (
            f"บันทึกสำเร็จ: {product.name} "
            f"{'เพิ่ม' if action == MovementType.IN else 'ตัด'} {quantity} {product.unit} "
            f"คงเหลือ {after_stock} {product.unit}"
        ),
        [product],
        ChatAssistantAction(
            type=(
                "stock_in"
                if action == MovementType.IN
                else ("stock_issue" if action == MovementType.ISSUE else "stock_out")
            ),
            barcode=product.barcode,
            product_name=product.name,
            quantity=quantity,
            previous_stock=before_stock,
            current_stock=after_stock,
            low_stock=after_stock <= product.minimum_stock,
            movement_id=record.id,
        ),
        notification,
    )


def build_ai_chat_reply(message: str, matched_products: list[Product], summary: dict[str, Any]) -> str | None:
    api_key = os.getenv("OPENAI_API_KEY", "").strip()
    if not api_key:
        return None

    model = os.getenv("OPENAI_MODEL", "gpt-4.1-mini").strip() or "gpt-4.1-mini"
    inventory_lines = [summarize_product(item) for item in matched_products[:8]]
    if not inventory_lines:
        low_stock_items = summary["low_stock_items"][:8]
        inventory_lines = [summarize_product(item) for item in low_stock_items]

    system_prompt = (
        "You are a Thai inventory assistant. Answer briefly in Thai using only the inventory data provided. "
        "Do not invent products or quantities. If the data is insufficient, say so clearly."
    )
    user_prompt = (
        f"Question: {message}\n"
        f"Inventory summary: total_products={summary['total_products']}, total_units={summary['total_units']}, "
        f"low_stock_count={summary['low_stock_count']}\n"
        "Relevant inventory:\n"
        + ("\n".join(inventory_lines) if inventory_lines else "No relevant products found.")
    )
    payload = json.dumps(
        {
            "model": model,
            "input": [
                {"role": "system", "content": [{"type": "input_text", "text": system_prompt}]},
                {"role": "user", "content": [{"type": "input_text", "text": user_prompt}]},
            ],
            "max_output_tokens": 220,
        }
    ).encode("utf-8")
    request = urllib.request.Request(
        "https://api.openai.com/v1/responses",
        data=payload,
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=20) as response:
            body = json.loads(response.read().decode("utf-8"))
    except (urllib.error.URLError, TimeoutError, json.JSONDecodeError):
        return None

    output = body.get("output", [])
    parts: list[str] = []
    for item in output:
        for content in item.get("content", []):
            text = content.get("text")
            if text:
                parts.append(text.strip())
    combined = " ".join(part for part in parts if part).strip()
    return combined or None


def detect_export_request(message: str) -> str | None:
    normalized = normalize_chat_text(message)
    if not any(keyword in normalized for keyword in ["ไฟล์", "ดาวน์โหลด", "export", "โหลด", "ขอ"]):
        return None
    if "สินค้า" in normalized and "csv" in normalized:
        return "products_csv"
    if "ผู้ใช้" in normalized and "csv" in normalized:
        return "users_csv"
    if ("ประวัติ" in normalized or "movement" in normalized) and "csv" in normalized:
        return "movements_csv"
    if "excel" in normalized or "xlsx" in normalized:
        return "all_xlsx"
    if "csv" in normalized:
        if "สินค้า" in normalized:
            return "products_csv"
        if "ผู้ใช้" in normalized:
            return "users_csv"
        if "ประวัติ" in normalized or "movement" in normalized:
            return "movements_csv"
        return "products_csv"
    return None


def is_product_listing_request(normalized: str) -> bool:
    return any(
        keyword in normalized
        for keyword in [
            "มีสินค้าอะไรบ้าง",
            "มีสินค้าอะไรในระบบบ้าง",
            "ขอดูรายการสินค้า",
            "แสดงสินค้าทั้งหมด",
            "รายการสินค้า",
            "สินค้ามีอะไรบ้าง",
        ]
    )


def csv_response(filename: str, headers: list[str], rows: list[list[Any]]) -> Response:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(headers)
    writer.writerows(rows)
    csv_content = "\ufeff" + buffer.getvalue()
    return Response(
        content=csv_content,
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )


def auto_fit_worksheet_columns(worksheet: Any) -> None:
    for column_cells in worksheet.columns:
        values = ["" if cell.value is None else str(cell.value) for cell in column_cells]
        max_length = max((len(value) for value in values), default=0)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(
            max(max_length + 2, 12),
            40,
        )


def ensure_product_for_scan(payload: ScanRequest) -> tuple[Product, bool, GoogleSheetProductCreateResult | None]:
    existing = products.get(payload.barcode)
    if existing:
        return existing, False, None

    if not payload.auto_create_product:
        raise HTTPException(
            status_code=404,
            detail="Barcode not found. Enable auto_create_product and send product_name to create it.",
        )

    if not payload.product_name or not payload.product_name.strip():
        raise HTTPException(
            status_code=400,
            detail="product_name is required when auto_create_product is true.",
        )

    product = Product(
        barcode=payload.barcode,
        sku=payload.product_sku,
        name=payload.product_name.strip(),
        unit=payload.product_unit or "pcs",
        minimum_stock=payload.product_minimum_stock,
        current_stock=0,
        category=payload.product_category,
        location=payload.product_location,
    )
    products[payload.barcode] = product
    save_product(product)

    create_status: GoogleSheetProductCreateResult | None = None
    try:
        client = sheet_client()
        if client.status().configured:
            create_status = client.append_product(product)
    except HTTPException:
        create_status = None

    return product, True, create_status


init_database()
load_state_from_db()


@app.on_event("startup")
def startup_refresh_state() -> None:
    load_state_from_db()


@app.websocket("/ws/realtime")
async def realtime_socket(websocket: WebSocket, token: str | None = Query(None)) -> None:
    user = user_from_token(token)
    if not user:
        await websocket.close(code=1008, reason="Invalid or expired access token.")
        return

    await realtime_manager.connect(websocket)
    await websocket.send_json(
        RealtimeEvent(
            type="connected",
            timestamp=utc_now(),
            payload={"user_id": user.user_id},
        ).model_dump(mode="json")
    )
    try:
        while True:
            await websocket.receive_text()
    except WebSocketDisconnect:
        realtime_manager.disconnect(websocket)
    except Exception:
        realtime_manager.disconnect(websocket)


@app.get("/")
def root() -> dict:
    return {
        "message": "Stock Scanner API is running.",
        "mobile_ready": True,
        "features": [
            "barcode scan",
            "stock in",
            "stock out",
            "issue for usage",
            "movement audit log",
            "sqlite persistence",
            "bearer token auth",
            "google sheet sync",
            "users dropdown",
            "auto product creation",
        ],
    }


@app.get("/health")
def health() -> dict:
    return {
        "status": "ok",
        "timestamp": utc_now(),
        "features": {
            "assistant_chat": True,
            "ai_chat": ai_chat_enabled(),
        },
    }


@app.get("/products", response_model=list[Product])
def list_products(low_stock_only: bool = False) -> list[Product]:
    items = list(products.values())
    if low_stock_only:
        items = [item for item in items if item.current_stock <= item.minimum_stock]
    return sorted(items, key=lambda item: item.name.lower())


@app.get("/products/barcode/next", response_model=BarcodeSuggestion)
def next_product_barcode() -> BarcodeSuggestion:
    return BarcodeSuggestion(
        barcode=generate_next_internal_barcode(),
        format="code128",
        message="Generated the next internal barcode for a new product.",
    )


@app.get("/products/{barcode}", response_model=Product)
def get_product(barcode: str) -> Product:
    product = products.get(barcode)
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    return product


@app.post("/products/upsert", response_model=Product)
async def upsert_product(
    payload: ProductUpsert,
    request: Request,
    requester_id: str | None = Query(None, min_length=1),
) -> Product:
    require_admin_request(request, requester_id)
    existing = products.get(payload.barcode)
    current_stock = existing.current_stock if existing else payload.opening_balance

    product = Product(
        barcode=payload.barcode,
        sku=payload.sku,
        name=payload.name,
        unit=payload.unit,
        minimum_stock=payload.minimum_stock,
        current_stock=current_stock,
        category=payload.category,
        location=payload.location,
    )
    products[payload.barcode] = product
    save_product(product)
    await broadcast_realtime_event(
        "product_updated",
        {
            "barcode": product.barcode,
            "product": product.model_dump(mode="json"),
        },
    )
    return product


@app.post("/auth/login", response_model=LoginResponse)
def login(payload: LoginRequest) -> LoginResponse:
    user = users.get(payload.user_id)
    if not user or not verify_pin(payload.pin, user.pin_hash):
        raise HTTPException(status_code=401, detail="Invalid user id or PIN.")
    if not user.active:
        raise HTTPException(status_code=403, detail="User is inactive.")
    access_token, expires_at = create_session(user.user_id)
    return LoginResponse(
        access_token=access_token,
        expires_at=expires_at,
        user=to_public_user(user),
    )


@app.get("/auth/me", response_model=UserPublic)
def auth_me(request: Request) -> UserPublic:
    return to_public_user(resolve_request_user(request))


@app.post("/auth/logout")
def logout(request: Request) -> dict:
    token = bearer_token_from_request(request)
    if not token:
        raise HTTPException(status_code=401, detail="Authentication required.")
    delete_session(token)
    return {"status": "ok", "message": "Logged out successfully."}


@app.post("/auth/change-pin", response_model=StatusMessage)
def change_pin(payload: ChangePinRequest, request: Request) -> StatusMessage:
    current_user = resolve_request_user(request)
    if not verify_pin(payload.current_pin, current_user.pin_hash):
        raise HTTPException(status_code=400, detail="Current PIN is incorrect.")
    if payload.current_pin == payload.new_pin:
        raise HTTPException(status_code=400, detail="The new PIN must be different from the current PIN.")

    updated_user = current_user.model_copy(update={"pin_hash": hash_pin(payload.new_pin)})
    users[updated_user.user_id] = updated_user
    save_user(updated_user)

    try:
        client = sheet_client()
        if client.status().configured:
            client.upsert_user(updated_user)
    except Exception as exc:
        logger.warning("Failed to sync updated PIN for user %s to Google Sheets: %s", updated_user.user_id, exc)

    return StatusMessage(message="PIN changed successfully.")


@app.patch("/auth/profile", response_model=UserPublic)
async def update_own_profile(payload: ProfileUpdateRequest, request: Request) -> UserPublic:
    current_user = resolve_request_user(request)
    updated_name = payload.user_name.strip()
    if not updated_name:
        raise HTTPException(status_code=400, detail="User name is required.")

    updated_user = current_user.model_copy(update={"user_name": updated_name})
    users[updated_user.user_id] = updated_user
    save_user(updated_user)

    try:
        client = sheet_client()
        if client.status().configured:
            client.upsert_user(updated_user)
    except Exception as exc:
        logger.warning("Failed to sync updated profile for user %s to Google Sheets: %s", updated_user.user_id, exc)

    public_user = to_public_user(updated_user)
    await broadcast_realtime_event(
        "user_updated",
        {
            "user_id": updated_user.user_id,
            "user": public_user.model_dump(mode="json"),
        },
    )
    return public_user


@app.get("/users", response_model=list[UserPublic])
def list_users(active_only: bool = True) -> list[UserPublic]:
    items = list(users.values())
    if active_only:
        items = [item for item in items if item.active]
    return [to_public_user(item) for item in sorted(items, key=lambda item: item.user_name.lower())]


@app.post("/users/upsert", response_model=UserPublic)
async def upsert_user(payload: UserUpsert, request: Request) -> UserPublic:
    require_admin_request(request, payload.requester_id)
    existing = users.get(payload.user_id)
    ensure_admin_account_retained(payload.user_id, payload.role, payload.active)
    pin_hash = hash_pin(payload.pin) if payload.pin else (existing.pin_hash if existing else None)
    if not pin_hash:
        raise HTTPException(status_code=400, detail="PIN is required for new users.")

    user = User(
        user_id=payload.user_id,
        user_name=payload.user_name,
        role=payload.role,
        active=payload.active,
        pin_hash=pin_hash,
        profile_image_url=payload.profile_image_url or (existing.profile_image_url if existing else None),
    )
    users[payload.user_id] = user
    save_user(user)

    try:
        client = sheet_client()
        if client.status().configured:
            client.upsert_user(user)
    except Exception as exc:
        logger.warning("Failed to sync user %s to Google Sheets: %s", user.user_id, exc)

    public_user = to_public_user(user)
    await broadcast_realtime_event(
        "user_updated",
        {
            "user_id": user.user_id,
            "user": public_user.model_dump(mode="json"),
        },
    )
    return public_user


@app.delete("/users/{user_id}", response_model=UserDeleteResponse)
async def delete_user(
    user_id: str,
    request: Request,
    requester_id: str | None = Query(None, min_length=1),
    delete_movements: bool = False,
) -> UserDeleteResponse:
    requester = require_admin_request(request, requester_id)
    target_user = users.get(user_id)
    if not target_user:
        raise HTTPException(status_code=404, detail=f"User {user_id} not found.")
    if requester.user_id == user_id:
        raise HTTPException(status_code=400, detail="You cannot delete the account currently in use.")

    ensure_admin_account_retained(user_id, target_user.role, False)

    profile_image_url = target_user.profile_image_url or ""
    if profile_image_url.startswith("/uploads/"):
        relative_path = profile_image_url.removeprefix("/uploads/")
        image_path = UPLOADS_DIR / relative_path
        if image_path.exists() and image_path.is_file():
            image_path.unlink(missing_ok=True)

    users.pop(user_id, None)
    deleted_movements = delete_user_record(user_id, delete_movements=delete_movements)
    if delete_movements:
        global movements
        movements = [record for record in movements if record.actor_id != user_id]

    await broadcast_realtime_event(
        "user_deleted",
        {
            "user_id": user_id,
            "delete_movements": delete_movements,
            "deleted_movements": deleted_movements,
        },
    )
    return UserDeleteResponse(
        message="User deleted successfully.",
        deleted_user_id=user_id,
        deleted_movements=deleted_movements,
    )


@app.post("/users/upload-profile-image", response_model=UserPublic)
async def upload_profile_image(
    request: Request,
    requester_id: str = Form(..., min_length=1),
    target_user_id: str = Form(..., min_length=1),
    image: UploadFile = File(...),
) -> UserPublic:
    can_manage_user_profile(resolve_request_user(request, requester_id).user_id, target_user_id)
    user = users.get(target_user_id)
    if not user:
        raise HTTPException(status_code=404, detail=f"User {target_user_id} not found.")

    extension = Path(image.filename or "").suffix.lower()
    if extension not in {".jpg", ".jpeg", ".png", ".webp"}:
        raise HTTPException(status_code=400, detail="Supported file types: jpg, jpeg, png, webp.")

    filename = f"{target_user_id.lower()}-{uuid4().hex}{extension}"
    destination = PROFILE_UPLOADS_DIR / filename
    content = await image.read()
    if not content:
        raise HTTPException(status_code=400, detail="Uploaded image is empty.")
    destination.write_bytes(content)

    user.profile_image_url = build_profile_image_url(filename)
    users[target_user_id] = user
    save_user(user)
    try:
        client = sheet_client()
        if client.status().configured:
            client.upsert_user(user)
    except Exception as exc:
        logger.warning("Failed to sync profile image for %s: %s", user.user_id, exc)
    public_user = to_public_user(user)
    await broadcast_realtime_event(
        "user_updated",
        {
            "user_id": user.user_id,
            "user": public_user.model_dump(mode="json"),
        },
    )
    return public_user


@app.post("/scan")
async def scan_item(payload: ScanRequest, request: Request) -> dict:
    request_user = resolve_request_user(request, payload.actor_id)
    if request.headers.get("Authorization") and request_user.user_id != payload.actor_id:
        raise HTTPException(status_code=403, detail="Authenticated user does not match actor_id.")
    user = get_user_or_404(payload.actor_id)
    payload.actor_name = user.user_name

    product, product_created, product_create_status = ensure_product_for_scan(payload)

    before_stock = product.current_stock
    delta = payload.quantity if payload.action == MovementType.IN else -payload.quantity
    after_stock = before_stock + delta

    if after_stock < 0:
        raise HTTPException(
            status_code=400,
            detail=f"Insufficient stock. Available={before_stock}, requested={payload.quantity}",
        )

    product.current_stock = after_stock
    products[payload.barcode] = product
    save_product(product)

    record = MovementRecord(
        id=str(uuid4()),
        barcode=product.barcode,
        product_name=product.name,
        action=payload.action,
        quantity=payload.quantity,
        before_stock=before_stock,
        after_stock=after_stock,
        actor_id=payload.actor_id,
        actor_name=payload.actor_name,
        note=payload.note,
        reference=payload.reference,
        created_at=utc_now(),
    )
    movements.insert(0, record)
    save_movement(record)

    append_status: GoogleSheetAppendResult | None = None
    stock_update_status: GoogleSheetStockUpdateResult | None = None
    google_sheets_error: str | None = None

    try:
        client = sheet_client()
        if client.status().configured:
            if product_created and not product_create_status:
                product_create_status = client.append_product(product)
            append_status = client.append_movement(record)
            stock_update_status = client.update_product_stock(product)
    except HTTPException as exc:
        google_sheets_error = str(exc.detail)
    except Exception as exc:
        google_sheets_error = str(exc)

    notification = create_notification(record)
    await broadcast_realtime_event(
        "stock_changed",
        {
            "barcode": product.barcode,
            "product": product.model_dump(mode="json"),
            "movement": record.model_dump(mode="json"),
            "notification": notification,
            "product_created": product_created,
            "low_stock": after_stock <= product.minimum_stock,
        },
    )

    return {
        "status": "ok",
        "product": product,
        "product_created": product_created,
        "product_create_sheet_status": (
            product_create_status.model_dump() if product_create_status else None
        ),
        "movement": record,
        "notification": notification,
        "low_stock": after_stock <= product.minimum_stock,
        "google_sheets": append_status.model_dump() if append_status else None,
        "google_sheets_stock_update": (
            stock_update_status.model_dump() if stock_update_status else None
        ),
        "google_sheets_error": google_sheets_error,
    }


@app.get("/movements", response_model=list[MovementRecord])
def list_movements(
    limit: int = Query(20, ge=1, le=200),
    barcode: str | None = None,
    actor_id: str | None = None,
) -> list[MovementRecord]:
    result = movements
    if barcode:
        result = [item for item in result if item.barcode == barcode]
    if actor_id:
        result = [item for item in result if item.actor_id == actor_id]
    return result[:limit]


@app.get("/notifications")
def list_notifications(limit: int = Query(20, ge=1, le=100)) -> list[dict]:
    return [create_notification(item) for item in movements[:limit]]


@app.get("/orders", response_model=list[Order])
def list_orders(
    request: Request,
    assigned_only: bool = False,
    mine_only: bool = False,
) -> list[Order]:
    user = resolve_request_user(request)
    result = list(orders)
    if not user.role.strip().lower() == "admin":
        result = [
            item
            for item in result
            if item.created_by_id == user.user_id or item.assigned_to_id == user.user_id
        ]
    if assigned_only:
        result = [item for item in result if item.assigned_to_id == user.user_id]
    if mine_only:
        result = [item for item in result if item.created_by_id == user.user_id]
    return result


@app.post("/orders", response_model=Order)
async def create_order(payload: OrderCreateRequest, request: Request) -> Order:
    user = resolve_request_user(request)
    assigned_user: User | None = None
    if payload.assigned_to_id:
        assigned_user = get_user_or_404(payload.assigned_to_id)

    built_items: list[OrderItem] = []
    for item in payload.items:
        product = products.get(item.barcode)
        if not product:
            raise HTTPException(status_code=404, detail=f"Product {item.barcode} not found.")
        built_items.append(
            OrderItem(
                barcode=product.barcode,
                product_name=product.name,
                quantity=item.quantity,
                unit=product.unit,
            )
        )

    now = utc_now()
    order = Order(
        id=str(uuid4()),
        customer_name=payload.customer_name.strip(),
        customer_phone=(payload.customer_phone or "").strip() or None,
        customer_address=(payload.customer_address or "").strip() or None,
        note=(payload.note or "").strip() or None,
        status=OrderStatus.ASSIGNED if assigned_user else OrderStatus.NEW,
        created_by_id=user.user_id,
        created_by_name=user.user_name,
        assigned_to_id=assigned_user.user_id if assigned_user else None,
        assigned_to_name=assigned_user.user_name if assigned_user else None,
        items=built_items,
        created_at=now,
        updated_at=now,
    )
    orders.insert(0, order)
    save_order(order)
    await broadcast_realtime_event(
        "order_created",
        {"order": order.model_dump(mode="json")},
    )
    return order


@app.post("/orders/{order_id}/assign", response_model=Order)
async def assign_order(order_id: str, payload: OrderAssignRequest, request: Request) -> Order:
    user = resolve_request_user(request)
    if user.role.strip().lower() not in {"admin", "staff"}:
        raise HTTPException(status_code=403, detail="Permission denied.")
    order = get_order_or_404(order_id)
    assigned_user = get_user_or_404(payload.assigned_to_id)
    updated = order.model_copy(
        update={
            "assigned_to_id": assigned_user.user_id,
            "assigned_to_name": assigned_user.user_name,
            "status": OrderStatus.ASSIGNED,
            "updated_at": utc_now(),
        }
    )
    for index, item in enumerate(orders):
        if item.id == order_id:
            orders[index] = updated
            break
    save_order(updated)
    await broadcast_realtime_event(
        "order_updated",
        {"order": updated.model_dump(mode="json")},
    )
    return updated


@app.post("/orders/{order_id}/status", response_model=Order)
async def update_order_status(order_id: str, payload: OrderStatusUpdateRequest, request: Request) -> Order:
    user = resolve_request_user(request)
    order = get_order_or_404(order_id)
    if user.role.strip().lower() != "admin" and order.assigned_to_id != user.user_id and order.created_by_id != user.user_id:
        raise HTTPException(status_code=403, detail="Permission denied.")
    updated = order.model_copy(update={"status": payload.status, "updated_at": utc_now()})
    for index, item in enumerate(orders):
        if item.id == order_id:
            orders[index] = updated
            break
    save_order(updated)
    await broadcast_realtime_event(
        "order_updated",
        {"order": updated.model_dump(mode="json")},
    )
    return updated


@app.get("/stock/summary")
def stock_summary() -> dict:
    return build_stock_summary_payload()


@app.post("/assistant/chat", response_model=ChatAssistantResponse)
async def assistant_chat(payload: ChatAssistantRequest, request: Request) -> ChatAssistantResponse:
    user = resolve_request_user(request)
    export_name = detect_export_request(payload.message)
    if export_name:
        params = {"movement_limit": 5000}
        token, expires_at = create_export_token(user.user_id, export_name, params)
        export_labels = {
            "all_xlsx": "ไฟล์ Excel",
            "products_csv": "ไฟล์สินค้า CSV",
            "users_csv": "ไฟล์ผู้ใช้ CSV",
            "movements_csv": "ไฟล์ประวัติ CSV",
        }
        return ChatAssistantResponse(
            message=f"นี่คือลิงก์{export_labels.get(export_name, 'ไฟล์ที่ขอ')} กดเปิดหรือคัดลอกลิงก์ได้เลย",
            download_link=ExportLinkResponse(
                url=f"/exports/download/{token}",
                expires_at=expires_at,
            ),
            ai_enabled=ai_chat_enabled(),
            used_ai=False,
        )

    action_message, action_products, action_result, notification = execute_chat_stock_action(
        payload.message,
        user,
    )
    if action_message:
        if action_result and notification:
            await broadcast_realtime_event(
                "stock_changed",
                {
                    "barcode": action_result.barcode,
                    "product": action_products[0].model_dump(mode="json"),
                    "movement": movements[0].model_dump(mode="json"),
                    "notification": notification,
                    "product_created": False,
                    "low_stock": action_result.low_stock,
                },
            )
        return ChatAssistantResponse(
            message=action_message,
            matched_products=action_products,
            action=action_result,
            ai_enabled=ai_chat_enabled(),
            used_ai=False,
        )

    summary = build_stock_summary_payload()
    normalized = normalize_chat_text(payload.message)

    if any(keyword in normalized for keyword in ["ใกล้หมด", "ของหมด", "สตอกต่ำ", "สต๊อกต่ำ", "lowstock", "หมดหรือยัง", "หมดยัง", "ต่ำกว่าขั้นต่ำ"]):
        low_stock_items = sorted(
            summary["low_stock_items"],
            key=lambda item: item.current_stock,
        )
        if not low_stock_items:
            return ChatAssistantResponse(
                message="ตอนนี้ยังไม่มีสินค้าที่อยู่ในระดับใกล้หมด",
                ai_enabled=ai_chat_enabled(),
            )
        preview = ", ".join(
            f"{item.name} ({item.current_stock}/{item.minimum_stock} {item.unit})"
            for item in low_stock_items[:5]
        )
        return ChatAssistantResponse(
            message=f"มีสินค้าใกล้หมด {len(low_stock_items)} รายการ: {preview}",
            matched_products=low_stock_items[:5],
            ai_enabled=ai_chat_enabled(),
        )

    if any(keyword in normalized for keyword in ["ทั้งหมด", "รวม", "กี่รายการ", "summary", "ภาพรวม", "สรุป"]):
        return ChatAssistantResponse(
            message=(
                f"ตอนนี้มีสินค้า {summary['total_products']} รายการ "
                f"รวมจำนวนคงเหลือ {summary['total_units']} ชิ้น/หน่วย "
                f"และมีสินค้าใกล้หมด {summary['low_stock_count']} รายการ"
            ),
            ai_enabled=ai_chat_enabled(),
        )

    if is_product_listing_request(normalized):
        product_list = sorted(products.values(), key=lambda item: item.name.lower())
        preview_items = product_list[:10]
        preview = ", ".join(
            f"{item.name} ({item.current_stock} {item.unit})"
            for item in preview_items
        )
        suffix = "" if len(product_list) <= 10 else " และยังมีรายการอื่นอีก"
        return ChatAssistantResponse(
            message=(
                f"ตอนนี้มีสินค้าในระบบ {len(product_list)} รายการ เช่น {preview}{suffix}"
            ),
            matched_products=preview_items,
            ai_enabled=ai_chat_enabled(),
        )

    matches = find_chat_products(payload.message)
    if len(matches) == 1:
        product = matches[0]
        status = "สินค้าใกล้หมดแล้ว" if product.current_stock <= product.minimum_stock else "สินค้ายังไม่ใกล้หมด"
        location = f" ตำแหน่ง {product.location}" if product.location else ""
        if any(keyword in normalized for keyword in ["ขั้นต่ำ", "ขั้นต่ำเท่าไหร่", "min", "minimum"]):
            deterministic_reply = (
                f"{product.name} ตั้งค่าขั้นต่ำไว้ {product.minimum_stock} {product.unit} "
                f"และตอนนี้คงเหลือ {product.current_stock} {product.unit} {status}.{location}"
            )
        else:
            deterministic_reply = (
                f"{product.name} คงเหลือ {product.current_stock} {product.unit} "
                f"ขั้นต่ำ {product.minimum_stock} {product.unit} {status}.{location}"
            )
    elif len(matches) > 1:
        preview = ", ".join(item.name for item in matches[:5])
        deterministic_reply = (
            f"เจอหลายรายการที่ใกล้เคียงกัน: {preview} ลองพิมพ์ชื่อหรือบาร์โค้ดให้เฉพาะเจาะจงขึ้น"
        )
    else:
        deterministic_reply = (
            "ยังหาไม่เจอ ลองพิมพ์ชื่อสินค้าให้สั้นลง หรือใช้บาร์โค้ด/รหัสสินค้า เช่น \"น้ำดื่มเหลือเท่าไหร่\""
        )

    ai_reply = build_ai_chat_reply(payload.message, matches, summary)
    return ChatAssistantResponse(
        message=ai_reply or deterministic_reply,
        matched_products=matches[:5],
        ai_enabled=ai_chat_enabled(),
        used_ai=bool(ai_reply),
    )


def build_products_csv_response() -> Response:
    rows = [
        [
            product.barcode,
            product.sku or "",
            product.name,
            product.unit,
            product.current_stock,
            product.category or "",
            product.location or "",
        ]
        for product in sorted(products.values(), key=lambda item: item.name.lower())
    ]
    return csv_response(
        "products.csv",
        ["เธเธฒเธฃเนเนเธเนเธ”", "SKU", "เธเธทเนเธญเธชเธดเธเธเนเธฒ", "เธซเธเนเธงเธข", "เธเธณเธเธงเธเธเธเน€เธซเธฅเธทเธญ", "เธซเธกเธงเธ”เธซเธกเธนเน", "เธ•เธณเนเธซเธเนเธเธเธฑเธ”เน€เธเนเธ"],
        rows,
    )


def build_users_csv_response() -> Response:
    rows = [
        [
            user.user_id,
            user.user_name,
            user.role,
            "TRUE" if user.active else "FALSE",
            user.profile_image_url or "",
        ]
        for user in sorted(users.values(), key=lambda item: item.user_name.lower())
    ]
    return csv_response(
        "users.csv",
        ["เธฃเธซเธฑเธชเธเธนเนเนเธเน", "เธเธทเนเธญเธเธนเนเนเธเน", "เธชเธดเธ—เธเธดเน", "เนเธเนเธเธฒเธเธญเธขเธนเน", "เธฃเธนเธเนเธเธฃเนเธเธฅเน"],
        rows,
    )


def build_movements_csv_response(
    limit: int = 500,
    barcode: str | None = None,
    actor_id: str | None = None,
) -> Response:
    result = movements
    if barcode:
        result = [item for item in result if item.barcode == barcode]
    if actor_id:
        result = [item for item in result if item.actor_id == actor_id]

    rows = [
        [
            item.id,
            item.created_at.isoformat(),
            item.barcode,
            item.product_name,
            item.action.value,
            item.quantity,
            item.before_stock,
            item.after_stock,
            item.actor_id,
            item.actor_name,
            item.note or "",
            item.reference or "",
        ]
        for item in result[:limit]
    ]
    return csv_response(
        "movements.csv",
        [
            "เธฃเธซเธฑเธชเธฃเธฒเธขเธเธฒเธฃ",
            "เธงเธฑเธเน€เธงเธฅเธฒ",
            "เธเธฒเธฃเนเนเธเนเธ”",
            "เธเธทเนเธญเธชเธดเธเธเนเธฒ",
            "เธเธฃเธฐเน€เธ เธ—",
            "เธเธณเธเธงเธ",
            "เธชเธ•เนเธญเธเธเนเธญเธเธ—เธณเธฃเธฒเธขเธเธฒเธฃ",
            "เธชเธ•เนเธญเธเธซเธฅเธฑเธเธ—เธณเธฃเธฒเธขเธเธฒเธฃ",
            "เธฃเธซเธฑเธชเธเธนเนเธ—เธณเธฃเธฒเธขเธเธฒเธฃ",
            "เธเธทเนเธญเธเธนเนเธ—เธณเธฃเธฒเธขเธเธฒเธฃ",
            "เธซเธกเธฒเธขเน€เธซเธ•เธธ",
            "เน€เธฅเธเธญเนเธฒเธเธญเธดเธ",
        ],
        rows,
    )


def build_all_xlsx_response(movement_limit: int = 5000) -> Response:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise HTTPException(
            status_code=500,
            detail="Missing Excel export dependency. Install 'openpyxl'.",
        ) from exc

    workbook = Workbook()

    products_sheet = workbook.active
    products_sheet.title = "เธชเธดเธเธเนเธฒ"
    products_sheet.append(
        ["เธเธฒเธฃเนเนเธเนเธ”", "SKU", "เธเธทเนเธญเธชเธดเธเธเนเธฒ", "เธซเธเนเธงเธข", "เธเธณเธเธงเธเธเธเน€เธซเธฅเธทเธญ", "เธซเธกเธงเธ”เธซเธกเธนเน", "เธ•เธณเนเธซเธเนเธเธเธฑเธ”เน€เธเนเธ"]
    )
    for product in sorted(products.values(), key=lambda item: item.name.lower()):
        products_sheet.append(
            [
                product.barcode,
                product.sku or "",
                product.name,
                product.unit,
                product.current_stock,
                product.category or "",
                product.location or "",
            ]
        )
    auto_fit_worksheet_columns(products_sheet)

    users_sheet = workbook.create_sheet("เธเธนเนเนเธเน")
    users_sheet.append(["เธฃเธซเธฑเธชเธเธนเนเนเธเน", "เธเธทเนเธญเธเธนเนเนเธเน", "เธชเธดเธ—เธเธดเน", "เนเธเนเธเธฒเธเธญเธขเธนเน", "เธฃเธนเธเนเธเธฃเนเธเธฅเน"])
    for user in sorted(users.values(), key=lambda item: item.user_name.lower()):
        users_sheet.append(
            [user.user_id, user.user_name, user.role, user.active, user.profile_image_url or ""]
        )
    auto_fit_worksheet_columns(users_sheet)

    movements_sheet = workbook.create_sheet("เธเธฃเธฐเธงเธฑเธ•เธด")
    movements_sheet.append(
        [
            "เธฃเธซเธฑเธชเธฃเธฒเธขเธเธฒเธฃ",
            "เธงเธฑเธเน€เธงเธฅเธฒ",
            "เธเธฒเธฃเนเนเธเนเธ”",
            "เธเธทเนเธญเธชเธดเธเธเนเธฒ",
            "เธเธฃเธฐเน€เธ เธ—",
            "เธเธณเธเธงเธ",
            "เธชเธ•เนเธญเธเธเนเธญเธเธ—เธณเธฃเธฒเธขเธเธฒเธฃ",
            "เธชเธ•เนเธญเธเธซเธฅเธฑเธเธ—เธณเธฃเธฒเธขเธเธฒเธฃ",
            "เธฃเธซเธฑเธชเธเธนเนเธ—เธณเธฃเธฒเธขเธเธฒเธฃ",
            "เธเธทเนเธญเธเธนเนเธ—เธณเธฃเธฒเธขเธเธฒเธฃ",
            "เธซเธกเธฒเธขเน€เธซเธ•เธธ",
            "เน€เธฅเธเธญเนเธฒเธเธญเธดเธ",
        ]
    )
    for item in movements[:movement_limit]:
        movements_sheet.append(
            [
                item.id,
                item.created_at.isoformat(),
                item.barcode,
                item.product_name,
                item.action.value,
                item.quantity,
                item.before_stock,
                item.after_stock,
                item.actor_id,
                item.actor_name,
                item.note or "",
                item.reference or "",
            ]
        )
    auto_fit_worksheet_columns(movements_sheet)

    output = io.BytesIO()
    workbook.save(output)
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="stock-data.xlsx"',
        },
    )


@app.post("/exports/link", response_model=ExportLinkResponse)
def create_export_link(
    payload: ExportLinkRequest,
    request: Request,
    requester_id: str | None = Query(None, min_length=1),
) -> ExportLinkResponse:
    user = require_admin_request(request, requester_id)
    params = {
        "movement_limit": payload.movement_limit,
        "barcode": payload.barcode,
        "actor_id": payload.actor_id,
    }
    token, expires_at = create_export_token(user.user_id, payload.export_name, params)
    return ExportLinkResponse(
        url=f"/exports/download/{token}",
        expires_at=expires_at,
    )


@app.get("/exports/download/{token}")
def download_export_by_token(token: str) -> Response:
    export_name, params, _user = consume_export_token(token)
    if export_name == "products_csv":
        rows = [
            [
                product.barcode,
                product.sku or "",
                product.name,
                product.unit,
                product.current_stock,
                product.category or "",
                product.location or "",
            ]
            for product in sorted(products.values(), key=lambda item: item.name.lower())
        ]
        return csv_response(
            "products.csv",
            ["บาร์โค้ด", "SKU", "ชื่อสินค้า", "หน่วย", "จำนวนคงเหลือ", "หมวดหมู่", "ตำแหน่งจัดเก็บ"],
            rows,
        )
    if export_name == "users_csv":
        rows = [
            [
                user.user_id,
                user.user_name,
                user.role,
                "TRUE" if user.active else "FALSE",
                user.profile_image_url or "",
            ]
            for user in sorted(users.values(), key=lambda item: item.user_name.lower())
        ]
        return csv_response(
            "users.csv",
            ["รหัสผู้ใช้", "ชื่อผู้ใช้", "สิทธิ์", "ใช้งานอยู่", "รูปโปรไฟล์"],
            rows,
        )
    if export_name == "movements_csv":
        movement_limit = int(params.get("movement_limit") or 500)
        result = movements
        barcode = params.get("barcode")
        actor_id = params.get("actor_id")
        if barcode:
            result = [item for item in result if item.barcode == barcode]
        if actor_id:
            result = [item for item in result if item.actor_id == actor_id]
        rows = [
            [
                item.id,
                item.created_at.isoformat(),
                item.barcode,
                item.product_name,
                item.action.value,
                item.quantity,
                item.before_stock,
                item.after_stock,
                item.actor_id,
                item.actor_name,
                item.note or "",
                item.reference or "",
            ]
            for item in result[: min(max(movement_limit, 1), 5000)]
        ]
        return csv_response(
            "movements.csv",
            [
                "รหัสรายการ",
                "วันเวลา",
                "บาร์โค้ด",
                "ชื่อสินค้า",
                "ประเภท",
                "จำนวน",
                "สต๊อกก่อนทำรายการ",
                "สต๊อกหลังทำรายการ",
                "รหัสผู้ทำรายการ",
                "ชื่อผู้ทำรายการ",
                "หมายเหตุ",
                "เลขอ้างอิง",
            ],
            rows,
        )
    if export_name == "all_xlsx":
        movement_limit = int(params.get("movement_limit") or 5000)
        try:
            from openpyxl import Workbook
        except ImportError as exc:
            raise HTTPException(
                status_code=500,
                detail="Missing Excel export dependency. Install 'openpyxl'.",
            ) from exc

        workbook = Workbook()

        products_sheet = workbook.active
        products_sheet.title = "สินค้า"
        products_sheet.append(
            ["บาร์โค้ด", "SKU", "ชื่อสินค้า", "หน่วย", "จำนวนคงเหลือ", "หมวดหมู่", "ตำแหน่งจัดเก็บ"]
        )
        for product in sorted(products.values(), key=lambda item: item.name.lower()):
            products_sheet.append(
                [
                    product.barcode,
                    product.sku or "",
                    product.name,
                    product.unit,
                    product.current_stock,
                    product.category or "",
                    product.location or "",
                ]
            )
        auto_fit_worksheet_columns(products_sheet)

        users_sheet = workbook.create_sheet("ผู้ใช้")
        users_sheet.append(["รหัสผู้ใช้", "ชื่อผู้ใช้", "สิทธิ์", "ใช้งานอยู่", "รูปโปรไฟล์"])
        for user in sorted(users.values(), key=lambda item: item.user_name.lower()):
            users_sheet.append(
                [user.user_id, user.user_name, user.role, user.active, user.profile_image_url or ""]
            )
        auto_fit_worksheet_columns(users_sheet)

        movements_sheet = workbook.create_sheet("ประวัติ")
        movements_sheet.append(
            [
                "รหัสรายการ",
                "วันเวลา",
                "บาร์โค้ด",
                "ชื่อสินค้า",
                "ประเภท",
                "จำนวน",
                "สต๊อกก่อนทำรายการ",
                "สต๊อกหลังทำรายการ",
                "รหัสผู้ทำรายการ",
                "ชื่อผู้ทำรายการ",
                "หมายเหตุ",
                "เลขอ้างอิง",
            ]
        )
        for item in movements[: min(max(movement_limit, 1), 20000)]:
            movements_sheet.append(
                [
                    item.id,
                    item.created_at.isoformat(),
                    item.barcode,
                    item.product_name,
                    item.action.value,
                    item.quantity,
                    item.before_stock,
                    item.after_stock,
                    item.actor_id,
                    item.actor_name,
                    item.note or "",
                    item.reference or "",
                ]
            )
        auto_fit_worksheet_columns(movements_sheet)

        output = io.BytesIO()
        workbook.save(output)
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": 'attachment; filename="stock-data.xlsx"',
            },
        )
    raise HTTPException(status_code=400, detail="Unsupported export type.")


@app.get("/exports/products.csv")
def export_products_csv(
    request: Request,
    requester_id: str | None = Query(None, min_length=1),
) -> Response:
    require_admin_request(request, requester_id)
    rows = [
        [
            product.barcode,
            product.sku or "",
            product.name,
            product.unit,
            product.current_stock,
            product.category or "",
            product.location or "",
        ]
        for product in sorted(products.values(), key=lambda item: item.name.lower())
    ]
    return csv_response(
        "products.csv",
        ["บาร์โค้ด", "SKU", "ชื่อสินค้า", "หน่วย", "จำนวนคงเหลือ", "หมวดหมู่", "ตำแหน่งจัดเก็บ"],
        rows,
    )


@app.get("/exports/users.csv")
def export_users_csv(
    request: Request,
    requester_id: str | None = Query(None, min_length=1),
) -> Response:
    require_admin_request(request, requester_id)
    rows = [
        [
            user.user_id,
            user.user_name,
            user.role,
            "TRUE" if user.active else "FALSE",
            user.profile_image_url or "",
        ]
        for user in sorted(users.values(), key=lambda item: item.user_name.lower())
    ]
    return csv_response(
        "users.csv",
        ["รหัสผู้ใช้", "ชื่อผู้ใช้", "สิทธิ์", "ใช้งานอยู่", "รูปโปรไฟล์"],
        rows,
    )


@app.get("/exports/movements.csv")
def export_movements_csv(
    request: Request,
    limit: int = Query(500, ge=1, le=5000),
    barcode: str | None = None,
    actor_id: str | None = None,
    requester_id: str | None = Query(None, min_length=1),
) -> Response:
    require_admin_request(request, requester_id)
    result = movements
    if barcode:
        result = [item for item in result if item.barcode == barcode]
    if actor_id:
        result = [item for item in result if item.actor_id == actor_id]

    rows = [
        [
            item.id,
            item.created_at.isoformat(),
            item.barcode,
            item.product_name,
            item.action.value,
            item.quantity,
            item.before_stock,
            item.after_stock,
            item.actor_id,
            item.actor_name,
            item.note or "",
            item.reference or "",
        ]
        for item in result[:limit]
    ]
    return csv_response(
        "movements.csv",
        [
            "รหัสรายการ",
            "วันเวลา",
            "บาร์โค้ด",
            "ชื่อสินค้า",
            "ประเภท",
            "จำนวน",
            "สต็อกก่อนทำรายการ",
            "สต็อกหลังทำรายการ",
            "รหัสผู้ทำรายการ",
            "ชื่อผู้ทำรายการ",
            "หมายเหตุ",
            "เลขอ้างอิง",
        ],
        rows,
    )


@app.get("/exports/all.xlsx")
def export_all_xlsx(
    request: Request,
    movement_limit: int = Query(5000, ge=1, le=20000),
    requester_id: str | None = Query(None, min_length=1),
) -> Response:
    require_admin_request(request, requester_id)
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise HTTPException(
            status_code=500,
            detail="Missing Excel export dependency. Install 'openpyxl'.",
        ) from exc

    workbook = Workbook()

    products_sheet = workbook.active
    products_sheet.title = "สินค้า"
    products_sheet.append(
        ["บาร์โค้ด", "SKU", "ชื่อสินค้า", "หน่วย", "จำนวนคงเหลือ", "หมวดหมู่", "ตำแหน่งจัดเก็บ"]
    )
    for product in sorted(products.values(), key=lambda item: item.name.lower()):
        products_sheet.append(
            [
                product.barcode,
                product.sku or "",
                product.name,
                product.unit,
                product.current_stock,
                product.category or "",
                product.location or "",
            ]
        )
    auto_fit_worksheet_columns(products_sheet)

    users_sheet = workbook.create_sheet("ผู้ใช้")
    users_sheet.append(["รหัสผู้ใช้", "ชื่อผู้ใช้", "สิทธิ์", "ใช้งานอยู่", "รูปโปรไฟล์"])
    for user in sorted(users.values(), key=lambda item: item.user_name.lower()):
        users_sheet.append(
            [user.user_id, user.user_name, user.role, user.active, user.profile_image_url or ""]
        )
    auto_fit_worksheet_columns(users_sheet)

    movements_sheet = workbook.create_sheet("ประวัติ")
    movements_sheet.append(
        [
            "รหัสรายการ",
            "วันเวลา",
            "บาร์โค้ด",
            "ชื่อสินค้า",
            "ประเภท",
            "จำนวน",
            "สต็อกก่อนทำรายการ",
            "สต็อกหลังทำรายการ",
            "รหัสผู้ทำรายการ",
            "ชื่อผู้ทำรายการ",
            "หมายเหตุ",
            "เลขอ้างอิง",
        ]
    )
    for item in movements[:movement_limit]:
        movements_sheet.append(
            [
                item.id,
                item.created_at.isoformat(),
                item.barcode,
                item.product_name,
                item.action.value,
                item.quantity,
                item.before_stock,
                item.after_stock,
                item.actor_id,
                item.actor_name,
                item.note or "",
                item.reference or "",
            ]
        )
    auto_fit_worksheet_columns(movements_sheet)

    output = io.BytesIO()
    workbook.save(output)
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="stock-data.xlsx"',
        },
    )


@app.get("/integrations/google-sheets/bootstrap", response_model=GoogleSheetBootstrap)
def google_sheets_bootstrap() -> GoogleSheetBootstrap:
    return GoogleSheetBootstrap(
        status="ready",
        message=(
            "Use your existing Google Sheet as the source of truth for products and users, "
            "then append every scan transaction into a movements sheet."
        ),
        recommended_tabs=["products", "movements", "users"],
        expected_product_columns=[
            "barcode",
            "sku",
            "name",
            "unit",
            "current_stock",
            "minimum_stock",
            "category",
            "location",
        ],
        expected_movement_columns=[
            "movement_id",
            "timestamp",
            "barcode",
            "product_name",
            "action",
            "quantity",
            "before_stock",
            "after_stock",
            "actor_id",
            "actor_name",
            "note",
            "reference",
        ],
        expected_user_columns=[
            "user_id",
            "user_name",
            "role",
            "active",
            "pin",
            "profile_image_url",
        ],
        mapping=sheet_mapping,
        environment_keys=[
            "GOOGLE_SHEETS_SPREADSHEET_ID",
            "GOOGLE_SERVICE_ACCOUNT_FILE",
        ],
    )


@app.get("/integrations/google-sheets/status", response_model=GoogleSheetStatus)
def google_sheets_status() -> GoogleSheetStatus:
    return sheet_client().status()


@app.post("/integrations/google-sheets/config", response_model=GoogleSheetStatus)
def configure_google_sheets(
    payload: GoogleSheetConfigRequest,
    request: Request,
    requester_id: str | None = Query(None, min_length=1),
) -> GoogleSheetStatus:
    require_admin_request(request, requester_id)
    global sheet_mapping
    sheet_mapping = SheetMapping(**payload.model_dump())
    return sheet_client().status()


@app.post("/integrations/google-sheets/sync/products", response_model=GoogleSheetSyncResult)
async def sync_products_from_google_sheets(
    request: Request,
    requester_id: str | None = Query(None, min_length=1),
) -> GoogleSheetSyncResult:
    require_admin_request(request, requester_id)
    result = sheet_client().read_products()
    await broadcast_realtime_event(
        "products_synced",
        {
            "imported_count": result.imported_count,
            "skipped_rows": result.skipped_rows,
        },
    )
    return result


@app.post("/integrations/google-sheets/sync/users", response_model=GoogleSheetSyncResult)
async def sync_users_from_google_sheets(
    request: Request,
    requester_id: str | None = Query(None, min_length=1),
) -> GoogleSheetSyncResult:
    require_admin_request(request, requester_id)
    result = sheet_client().read_users()
    await broadcast_realtime_event(
        "users_synced",
        {
            "imported_count": result.imported_count,
            "skipped_rows": result.skipped_rows,
        },
    )
    return result


@app.post("/integrations/google-sheets/append-test", response_model=GoogleSheetAppendResult)
def append_test_row(
    request: Request,
    requester_id: str | None = Query(None, min_length=1),
) -> GoogleSheetAppendResult:
    require_admin_request(request, requester_id)
    test_record = MovementRecord(
        id=str(uuid4()),
        barcode="TEST-BARCODE",
        product_name="Connection Test",
        action=MovementType.IN,
        quantity=1,
        before_stock=0,
        after_stock=1,
        actor_id="system",
        actor_name="System Test",
        note="Google Sheets append test",
        reference="connectivity-check",
        created_at=utc_now(),
    )
    return sheet_client().append_movement(test_record)


@app.post(
    "/integrations/google-sheets/sync/stocks",
    response_model=GoogleSheetBulkSyncResult,
)
async def sync_stock_balances_to_google_sheets(
    request: Request,
    requester_id: str | None = Query(None, min_length=1),
) -> GoogleSheetBulkSyncResult:
    require_admin_request(request, requester_id)
    result = sheet_client().sync_all_product_stocks()
    await broadcast_realtime_event(
        "stocks_synced",
        {
            "updated_count": result.updated_count,
            "skipped_count": result.skipped_count,
        },
    )
    return result


@app.post("/webhook")
def webhook(data: dict, request: Request) -> dict:
    expected_secret = os.getenv("WEBHOOK_SECRET")
    if expected_secret:
        provided_secret = request.headers.get("X-Webhook-Secret")
        if provided_secret != expected_secret:
            raise HTTPException(status_code=401, detail="Invalid webhook secret.")
    return {"status": "ok", "received": data}
